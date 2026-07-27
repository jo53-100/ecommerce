"""
Importa el historico 2026 desde NOMINA_2026.xlsx y BASE_DE_DATOS_2026.xlsx.

Decisiones para NO duplicar ingresos:
  - Los ingresos se toman de las hojas por-servicio de BASE_DE_DATOS
    (granularidad diaria: dia x mes), que son las mas detalladas.
  - De NOMINA solo se siembran las EMPLEADAS y sus tasas; sus totales
    diarios de tarjeta/efectivo son el MISMO dinero contado de otra forma,
    asi que no se importan como ingreso (se duplicarian).
  - Los gastos se toman de la seccion GASTOS MENSUAL de la hoja GLOBAL.

Uso:
    python manage.py import_excel /ruta/NOMINA_2026.xlsx /ruta/BASE_DE_DATOS_2026.xlsx
"""
import datetime as dt
from decimal import Decimal
import openpyxl
from django.core.management.base import BaseCommand
from django.db import transaction as db_transaction
from clinic.models import Employee, ServiceType, Transaction, Expense
from store.models.products import Products

YEAR = 2026
# columna 2 (B) = enero ... columna 13 (M) = diciembre
COL_TO_MONTH = {c: c - 1 for c in range(2, 14)}

PRODUCT_SHEETS = {
    "ANTITRASPIRANTE", "PROTECTOR JUANETE", "DSINFECTANTE", "JABON DE AZUFRE",
    "UREA", "LEBEL", "BARNICES", "SEPARADORES DE DEDOS", "REXPLUX",
    "ESENCIAS MASAJE", "GUANTE PLANTAL", "PLANTILLAS",
}
TIP_SHEETS = {"PROPINAS"}
SKIP_SHEETS = {"GLOBAL"}

EXPENSE_MAP = {
    "SALARIOS": "salary", "RENTA": "rent", "LUZ": "utilities",
    "INTERNET": "utilities", "CELULAR": "utilities", "SAT": "tax",
    "CONTADOR": "tax", "COMPRAS VARIAS": "supplies",
}


def num(v):
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return Decimal(str(v))
    return None


def safe_date(month, day):
    try:
        return dt.date(YEAR, month, int(day))
    except (ValueError, TypeError):
        return None


class Command(BaseCommand):
    help = "Importa el historico 2026 de los Excel a la base de datos."

    def add_arguments(self, parser):
        parser.add_argument("nomina")
        parser.add_argument("base_datos")

    @db_transaction.atomic
    def handle(self, *args, **opts):
        self.seed_employees()
        self.import_services_and_products(opts["base_datos"])
        self.import_expenses(opts["base_datos"])
        self.summary()

    def seed_employees(self):
        # 10% del servicio sin IVA = 0.10 / 1.16 = 0.0862 (confirmado en NOMINA).
        data = [
            ("Laura", "admin", 3000, "0.0862"),
            ("Yara", "cosmetologa", 2000, "0"),     # tasa por confirmar con la duena
            ("Laura E.", "podologa", 1800, "0"),    # tasa por confirmar con la duena
        ]
        for name, role, base, rate in data:
            Employee.objects.get_or_create(
                name=name,
                defaults=dict(role=role, base_salary_weekly=base, commission_rate=Decimal(rate)),
            )
        self.stdout.write(f"Empleadas: {Employee.objects.count()}")

    def import_services_and_products(self, path):
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        n_tx = skipped = 0
        for title in wb.sheetnames:
            key = title.strip().upper()
            if key in SKIP_SHEETS:
                continue
            ws = wb[title]
            if key in TIP_SHEETS:
                kind, svc, prod = "tip", None, None
            elif key in PRODUCT_SHEETS:
                kind = "product_sale"
                # El catalogo vive en la tienda. Lo historico entra sin
                # control de existencias (el Excel no traia stock) y fuera
                # del escaparate; se activa a mano cuando haga falta.
                prod, _ = Products.objects.get_or_create(
                    name=title.strip().title(),
                    defaults={"track_inventory": False, "for_sale": False},
                )
                svc = None
            else:
                kind = "service"
                svc, _ = ServiceType.objects.get_or_create(name=title.strip().title())
                prod = None

            for row in ws.iter_rows(min_row=3):
                day = row[0].value
                for col, cell in enumerate(row[1:], start=2):
                    month = COL_TO_MONTH.get(col)
                    amount = num(cell.value)
                    if month is None or amount is None or amount <= 0:
                        continue
                    d = safe_date(month, day)
                    if d is None:
                        skipped += 1
                        continue
                    Transaction.objects.create(
                        date=d, kind=kind, amount=amount,
                        service_type=svc, product=prod,
                        is_imported_aggregate=True,
                        note=f"Importado de hoja '{title.strip()}'",
                    )
                    n_tx += 1
        wb.close()
        self.stdout.write(f"Movimientos importados: {n_tx} (fechas invalidas omitidas: {skipped})")

    def import_expenses(self, path):
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb["GLOBAL"]
        rows = list(ws.iter_rows())
        n = 0
        # Bloque 1: GASTOS MENSUAL (ojo: en el Excel dice 'MESUAL') -> renta, salarios, etc.
        n += self._import_expense_block(rows, marker="GASTOS MES", skip_header_rows=1,
                                         category_for=lambda key: EXPENSE_MAP.get(key, "other"))
        # Bloque 2: PRODUCTOS QUE COMPRO -> insumos/compras (guantes, gasas, cafe, etc.).
        # Antes NO se importaba: ese costo real nunca llegaba al sistema (quedaba solo
        # en el Excel), lo que inflaba la utilidad mostrada en el dashboard de Django.
        n += self._import_expense_block(rows, marker="QUE COMPRO", skip_header_rows=2,
                                         category_for=lambda key: "supplies")
        wb.close()
        self.stdout.write(f"Gastos importados: {n}")

    def _import_expense_block(self, rows, marker, skip_header_rows, category_for):
        start = None
        for i, row in enumerate(rows):
            a = (row[0].value or "")
            if isinstance(a, str) and marker in a.strip().upper():
                start = i + skip_header_rows
                break
        if start is None:
            return 0
        n = 0
        for row in rows[start:]:
            label = row[0].value
            if not isinstance(label, str):
                continue
            key = label.strip().upper()
            if key == "TOTAL":
                break
            cat = category_for(key)
            for col, cell in enumerate(row[1:], start=2):
                month = COL_TO_MONTH.get(col)
                amount = num(cell.value)
                if month is None or amount is None or amount <= 0:
                    continue
                d = safe_date(month, 15)
                Expense.objects.create(date=d, category=cat, amount=amount, note=label.strip())
                n += 1
        return n

    def summary(self):
        from django.db.models import Sum
        svc = Transaction.objects.filter(kind="service").aggregate(s=Sum("amount"))["s"] or 0
        prod = Transaction.objects.filter(kind="product_sale").aggregate(s=Sum("amount"))["s"] or 0
        tips = Transaction.objects.filter(kind="tip").aggregate(s=Sum("amount"))["s"] or 0
        exp = Expense.objects.aggregate(s=Sum("amount"))["s"] or 0
        self.stdout.write(self.style.SUCCESS(
            f"\nRESUMEN 2026:\n"
            f"  Servicios:  ${svc:,.2f}\n"
            f"  Productos:  ${prod:,.2f}\n"
            f"  Propinas:   ${tips:,.2f}\n"
            f"  Gastos:     ${exp:,.2f}\n"
            f"  Catalogos:  {ServiceType.objects.count()} servicios, {Products.objects.count()} productos"
        ))
